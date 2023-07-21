import os
from openpyxl import load_workbook, Workbook
import config
import tool

# 获取文件名列表


def find_score_points(sheet):
    points = tool.find_cell(sheet, config.search_score_title)
    points = [(x + 1, y - 1) for (x, y) in points]
    return points


def sort_info(student_info, header):
    student_info = {k: student_info[k] for k in header}
    return student_info


def write_in_file(student_info):
    # 检查文件是否存在
    if not os.path.exists('统计结果.xlsx'):
        # 如果文件不存在，创建一个新的文件并写入表头和第一行的数据
        wb = Workbook()
        ws = wb.active
        ws.append(config.header)  # 添加表头
        ws.append(list(student_info.values()))  # 添加第一行的数据
    else:
        # 如果文件存在，加载该文件并追加新的数据
        wb = load_workbook('统计结果.xlsx')
        ws = wb.active
        ws.append(list(student_info.values()))  # 添加新的数据
    # 保存修改
    wb.save("统计结果.xlsx")


def start():

    # 预处理file文件夹里的所有文件，如果是以xls结尾，则将他们转换为xlsx
    tool.multi_xls2xlsx(config.dir_name)
    # 获取所有文件名
    file_list = tool.get_file_list(config.dir_name)

    for file in file_list:
        wb = load_workbook(filename=file)
        sheet = wb[wb.sheetnames[0]]
        student_info = tool.find_attr(sheet, config.attr)
        points = find_score_points(sheet)
        lessons_list = tool.get_lessons_dict(sheet, points)
        ave = tool.calculate_average(student_info["姓名"],
                                     lessons_list,
                                     config.calculate_condition)
        student_info["平均分"] = str(ave)
        student_info = sort_info(student_info, config.header)
        write_in_file(student_info)
        print(f'已写入：{student_info["姓名"]}')
    print("任务已完成! 请查看《统计结果.xlsx》")


if __name__ == "__main__":
    start()
